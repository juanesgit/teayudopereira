from __future__ import annotations
import io
import json
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.census import CensusRecord
from app.models.user import User, UserRole
from app.services.auth import get_current_user, get_current_user_from_query

log = logging.getLogger(__name__)
router = APIRouter(prefix="/census", tags=["Census"])

ALLOWED_ROLES = (UserRole.admin, UserRole.coordinator, UserRole.volunteer)


def _require_census_access(user: User):
    if user.role not in ALLOWED_ROLES:
        raise HTTPException(403, "Sin permisos")


def _payload(r: CensusRecord, registered_by_name: str = None) -> dict:
    return {
        "id": r.id,
        "full_name": r.full_name,
        "document_number": r.document_number,
        "age": r.age,
        "gender": r.gender,
        "phone": r.phone,
        "whatsapp": r.whatsapp,
        "address": r.address,
        "neighborhood": r.neighborhood,
        "lat": r.lat,
        "lng": r.lng,
        "people_count": r.people_count,
        "children_count": r.children_count,
        "adults_count": getattr(r, "adults_count", 0),
        "elderly_count": r.elderly_count,
        "needs": json.loads(r.needs) if r.needs else [],
        "vulnerable": json.loads(r.vulnerable) if r.vulnerable else [],
        "shelter_status": r.shelter_status,
        "notes": r.notes,
        "is_attended": getattr(r, "is_attended", False),
        "attended_at": (r.attended_at.isoformat() + "Z") if getattr(r, "attended_at", None) else None,
        "registered_by": r.registered_by,
        "registered_by_name": registered_by_name,
        "created_at": r.created_at.isoformat() + "Z",
        "updated_at": r.updated_at.isoformat() + "Z",
    }


@router.get("")
async def list_census(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_census_access(current_user)
    result = await db.execute(
        select(CensusRecord)
        .where(CensusRecord.is_active == True)  # noqa
        .order_by(CensusRecord.created_at.desc())
    )
    records = result.scalars().all()

    # Cargar nombres de registradores
    out = []
    for r in records:
        name = None
        if r.registered_by:
            u = (await db.execute(select(User).where(User.id == r.registered_by))).scalar_one_or_none()
            if u:
                name = f"{u.full_name}{' ' + u.last_name if u.last_name else ''}"
        out.append(_payload(r, name))
    return out


@router.post("")
async def create_census(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_census_access(current_user)

    name = str(body.get("full_name", "")).strip()
    if not name:
        raise HTTPException(400, "El nombre es requerido")

    # Detección de duplicados por documento
    doc = str(body.get("document_number", "") or "").strip()
    if doc:
        existing = (await db.execute(
            select(CensusRecord)
            .where(CensusRecord.document_number == doc, CensusRecord.is_active == True)  # noqa
        )).scalar_one_or_none()
        if existing:
            raise HTTPException(409, f"Ya existe un registro con el documento {doc} (ID #{existing.id}: {existing.full_name})")

    needs = body.get("needs", [])
    vulnerable = body.get("vulnerable", [])

    r = CensusRecord(
        full_name=name,
        document_number=body.get("document_number"),
        age=body.get("age"),
        gender=body.get("gender"),
        phone=body.get("phone"),
        whatsapp=body.get("whatsapp"),
        address=body.get("address"),
        neighborhood=body.get("neighborhood"),
        lat=body.get("lat"),
        lng=body.get("lng"),
        people_count=int(body.get("people_count", 1)),
        children_count=int(body.get("children_count", 0)),
        adults_count=int(body.get("adults_count", 0)),
        elderly_count=int(body.get("elderly_count", 0)),
        needs=json.dumps(needs) if needs else None,
        vulnerable=json.dumps(vulnerable) if vulnerable else None,
        shelter_status=body.get("shelter_status"),
        notes=body.get("notes"),
        registered_by=current_user.id,
    )
    db.add(r)
    await db.commit()
    await db.refresh(r)
    return _payload(r, f"{current_user.full_name}{' ' + current_user.last_name if current_user.last_name else ''}")


@router.patch("/{record_id}")
async def update_census(
    record_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_census_access(current_user)

    r = (await db.execute(select(CensusRecord).where(CensusRecord.id == record_id))).scalar_one_or_none()
    if not r:
        raise HTTPException(404, "Registro no encontrado")

    allowed = {"full_name", "document_number", "age", "gender", "phone", "whatsapp",
               "address", "neighborhood", "lat", "lng", "people_count", "children_count",
               "adults_count", "elderly_count", "shelter_status", "notes"}
    vals: dict = {"updated_at": datetime.utcnow()}
    for k, v in body.items():
        if k in allowed:
            vals[k] = v

    if "needs" in body:
        vals["needs"] = json.dumps(body["needs"]) if body["needs"] else None
    if "vulnerable" in body:
        vals["vulnerable"] = json.dumps(body["vulnerable"]) if body["vulnerable"] else None

    await db.execute(update(CensusRecord).where(CensusRecord.id == record_id).values(**vals))
    await db.commit()
    return {"ok": True}


@router.patch("/{record_id}/attend")
async def toggle_attended(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Alterna el estado de atención de un registro."""
    _require_census_access(current_user)
    r = (await db.execute(select(CensusRecord).where(CensusRecord.id == record_id))).scalar_one_or_none()
    if not r:
        raise HTTPException(404, "Registro no encontrado")
    new_val = not getattr(r, "is_attended", False)
    await db.execute(
        update(CensusRecord).where(CensusRecord.id == record_id).values(
            is_attended=new_val,
            attended_at=datetime.utcnow() if new_val else None,
            updated_at=datetime.utcnow(),
        )
    )
    await db.commit()
    return {"ok": True, "is_attended": new_val}


# ── Exportación Excel ─────────────────────────────────────────────────────────

@router.get("/export")
async def export_census_xlsx(
    need: Optional[str] = None,
    shelter: Optional[str] = None,
    attended: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_query),
):
    """Exporta el censo de afectados como archivo Excel. Solo admins."""
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Solo administradores pueden exportar")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    result = await db.execute(
        select(CensusRecord)
        .where(CensusRecord.is_active == True)  # noqa
        .order_by(CensusRecord.created_at.desc())
    )
    all_records = result.scalars().all()

    # Aplicar filtros
    records = []
    for r in all_records:
        if need and need not in (json.loads(r.needs) if r.needs else []):
            continue
        if shelter and r.shelter_status != shelter:
            continue
        if attended == "true" and not getattr(r, "is_attended", False):
            continue
        if attended == "false" and getattr(r, "is_attended", False):
            continue
        if search:
            s = search.lower()
            if not any([
                s in (r.full_name or "").lower(),
                s in (r.document_number or "").lower(),
                s in (r.neighborhood or "").lower(),
                s in (r.address or "").lower(),
            ]):
                continue
        records.append(r)

    # Cargar nombres de registradores una sola vez
    user_ids = {r.registered_by for r in records if r.registered_by}
    users = {}
    for uid in user_ids:
        u = (await db.execute(select(User).where(User.id == uid))).scalar_one_or_none()
        if u:
            users[uid] = f"{u.full_name}{' ' + u.last_name if u.last_name else ''}".strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Censo Afectados"

    COLS = 19
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Encabezado principal
    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    t = ws["A1"]
    t.value = "Te Ayudo Pereira — Censo de Afectados"
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="DC2626")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(COLS)}2")
    d = ws["A2"]
    d.value = f"Generado el {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC  ·  {len(records)} registro(s)"
    d.font = Font(italic=True, size=9, color="6B7280")
    d.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Cabeceras
    headers = [
        "Nombre completo", "Documento", "Edad", "Género",
        "Teléfono", "WhatsApp", "Dirección", "Barrio",
        "Personas", "Menores", "Adultos", "Adultos mayores",
        "Alojamiento", "Necesidades", "Vulnerabilidad", "Notas",
        "Atendido", "Registrado por", "Fecha registro",
    ]
    COLS = len(headers)
    hfill = PatternFill("solid", fgColor="111827")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[3].height = 20

    # Datos
    alt_fill = PatternFill("solid", fgColor="F9FAFB")
    shelter_labels = {
        "own_home": "Casa propia", "renting": "Alquiler",
        "shelter": "Albergue", "relative": "Familiar",
        "street": "Calle", None: "",
    }
    need_labels = {
        "food": "Alimentos", "water": "Agua", "medical": "Médico",
        "shelter": "Alojamiento", "clothing": "Ropa",
        "psychological": "Apoyo psicológico", "legal": "Legal",
        "baby": "Bebé", "pet": "Mascotas",
    }

    for i, r in enumerate(records, start=4):
        vuln_labels = {
            "pregnant": "Embarazada", "disability": "Discapacidad",
            "chronic_illness": "Enfermedad crónica", "mental_health": "Salud mental",
        }
        needs_str = ", ".join(need_labels.get(n, n) for n in (json.loads(r.needs) if r.needs else []))
        vuln_str = ", ".join(vuln_labels.get(v, v) for v in (json.loads(r.vulnerable) if r.vulnerable else []))
        shelter = shelter_labels.get(r.shelter_status, r.shelter_status or "")
        vol_name = users.get(r.registered_by, "desconocido") if r.registered_by else "—"
        fecha = r.created_at.strftime("%d/%m/%Y %H:%M") if r.created_at else ""
        atendido = "Sí" if getattr(r, "is_attended", False) else "No"

        row_data = [
            r.full_name, r.document_number, r.age, r.gender,
            r.phone, r.whatsapp, r.address, r.neighborhood,
            r.people_count, r.children_count, getattr(r, "adults_count", 0), r.elderly_count,
            shelter, needs_str, vuln_str, r.notes,
            atendido, vol_name, fecha,
        ]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border
            if fill:
                c.fill = fill
        ws.row_dimensions[i].height = 16

    # Anchos de columna
    col_widths = [28, 16, 8, 12, 14, 14, 30, 18, 10, 10, 10, 14, 16, 32, 24, 28, 10, 22, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"censo_afectados_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/{record_id}")
async def delete_census(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_census_access(current_user)
    await db.execute(
        update(CensusRecord)
        .where(CensusRecord.id == record_id)
        .values(is_active=False, updated_at=datetime.utcnow())
    )
    await db.commit()
    return {"ok": True}
