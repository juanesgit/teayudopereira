"""
Endpoints de administración — solo rol admin.
"""
from __future__ import annotations

import csv
import io
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from pydantic import BaseModel
from typing import Optional, List

from app.database import get_db
from app.models.user import User, UserRole
from app.services.auth import get_current_user
from app.services import sms

router = APIRouter(prefix="/admin", tags=["Administración"])


def _require_admin(current_user: User = Depends(get_current_user)) -> User:
    """Solo admin del sistema."""
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Solo administradores del sistema pueden acceder")
    return current_user


def _require_staff(current_user: User = Depends(get_current_user)) -> User:
    """Coordinadores Y admin del sistema."""
    if current_user.role not in (UserRole.coordinator, UserRole.admin):
        raise HTTPException(status_code=403, detail="Se requiere rol coordinador o admin")
    return current_user


# ─── Schemas ─────────────────────────────────────────────────────────────────

class SmsSendRequest(BaseModel):
    phones: List[str]          # números directos
    message: str


class SmsBroadcastRequest(BaseModel):
    message: str
    target: str = "all"        # "all" | "volunteers" | "coordinators"


class SmsSendResponse(BaseModel):
    ok: bool
    recipients: int
    message: str


# ─── Endpoints ───────────────────────────────────────────────────────────────

@router.post("/sms/send", response_model=SmsSendResponse)
async def sms_send_direct(
    data: SmsSendRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(_require_staff),
):
    """Envía SMS a una lista de números específicos. Coordinadores y admin."""
    if not data.phones:
        raise HTTPException(status_code=400, detail="Debes indicar al menos un número")
    if not data.message.strip():
        raise HTTPException(status_code=400, detail="El mensaje no puede estar vacío")

    background_tasks.add_task(sms.send_sms, data.phones, data.message.strip())
    return SmsSendResponse(ok=True, recipients=len(data.phones), message="SMS en envío")


@router.post("/sms/broadcast", response_model=SmsSendResponse)
async def sms_broadcast(
    data: SmsBroadcastRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(_require_staff),
):
    """Envía SMS masivo. Coordinadores y admin."""
    if not data.message.strip():
        raise HTTPException(status_code=400, detail="El mensaje no puede estar vacío")

    query = select(User).where(User.is_active == True)  # noqa: E712
    if data.target == "volunteers":
        query = query.where(User.role == UserRole.volunteer)
    elif data.target == "coordinators":
        query = query.where(User.role == UserRole.coordinator)
    else:  # "all"
        query = query.where(User.role.in_([UserRole.volunteer, UserRole.coordinator]))

    result = await db.execute(query)
    users = result.scalars().all()
    phones = [u.phone for u in users if u.phone]

    if not phones:
        return SmsSendResponse(ok=False, recipients=0, message="Sin destinatarios con teléfono")

    background_tasks.add_task(sms.send_sms, phones, data.message.strip())
    return SmsSendResponse(ok=True, recipients=len(phones), message=f"Broadcast a {len(phones)} número(s) en envío")


@router.get("/sms/consent-export")
async def export_sms_consents(current_user: User = Depends(_require_admin)):
    """Exporta CSV con todos los usuarios que dieron consentimiento SMS."""
    from app.database import AsyncSessionLocal
    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(User).where(User.sms_consent_at != None).order_by(User.sms_consent_at)  # noqa: E711
        )
        users = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Nombre", "Apellido", "Teléfono", "Rol", "Barrio", "Fecha consentimiento"])
    for u in users:
        writer.writerow([
            u.full_name,
            u.last_name or "",
            u.phone,
            u.role.value,
            u.neighborhood or "",
            u.sms_consent_at.strftime("%Y-%m-%d %H:%M:%S") if u.sms_consent_at else "",
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=consentimientos_sms.csv"},
    )


@router.post("/sms/blast-excel")
async def sms_blast_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    message: str = Form(...),
    comuna: str = Form(""),          # vacío = todas las comunas
    preview: bool = Form(False),     # True = solo devuelve conteo, no envía
    current_user: User = Depends(_require_admin),
):
    """
    Carga un Excel con columnas CEL / TELEFONO_1 / NODOS_COMUNAS_nombre,
    extrae números móviles válidos, filtra por comuna si se indica,
    y dispara el SMS masivo via Inalambria.
    Con preview=True solo devuelve el conteo y lista de comunas sin enviar.
    """
    import re
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl no instalado")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "El archivo está vacío")

    # Detectar cabeceras
    headers = [str(h).strip().upper() if h else "" for h in rows[0]]

    def col(names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    idx_cel     = col(["CEL"])
    idx_tel1    = col(["TELEFONO_1", "TELEFONO1"])
    idx_comuna  = col(["NODOS_COMUNAS_NOMBRE", "NODOS_COMUNAS_nombre".upper(), "COMUNA"])
    idx_nombre  = col(["DATA CLIENTES_NOMBRE", "NOMBRE", "DATA CLIENTES_nombre".upper()])

    if idx_cel is None and idx_tel1 is None:
        raise HTTPException(400, "No se encontró columna CEL ni TELEFONO_1 en el archivo")

    def valid_mobile(v) -> str | None:
        if not v:
            return None
        digits = re.sub(r"\D", "", str(v))
        if len(digits) == 10 and digits.startswith("3"):
            return f"+57{digits}"
        if len(digits) == 12 and digits.startswith("573"):
            return f"+{digits}"
        return None

    comunas_set = set()
    phones = []
    seen = set()

    for row in rows[1:]:
        # Filtro de comuna
        row_comuna = str(row[idx_comuna]).strip() if idx_comuna is not None and row[idx_comuna] else ""
        if row_comuna:
            comunas_set.add(row_comuna)
        if comuna and row_comuna.lower() != comuna.lower():
            continue

        # Número: prioridad CEL, fallback TELEFONO_1
        raw = None
        if idx_cel is not None:
            raw = row[idx_cel]
        if not valid_mobile(raw) and idx_tel1 is not None:
            raw = row[idx_tel1]

        num = valid_mobile(raw)
        if num and num not in seen:
            seen.add(num)
            phones.append(num)

    wb.close()

    if preview:
        return {
            "total": len(phones),
            "comunas": sorted(comunas_set),
            "preview": True,
        }

    if not phones:
        raise HTTPException(400, "No se encontraron números móviles válidos con el filtro aplicado")

    if not message.strip():
        raise HTTPException(400, "El mensaje no puede estar vacío")

    background_tasks.add_task(sms.send_sms, phones, message.strip())
    return {
        "ok": True,
        "recipients": len(phones),
        "message": f"SMS en cola para {len(phones)} número(s). Filtro comuna: '{comuna or 'todas'}'.",
    }


@router.get("/sms/balance")
async def sms_balance(current_user: User = Depends(_require_admin)):
    """Consulta saldo de créditos SMS."""
    data = await sms.get_balance()
    if data is None:
        raise HTTPException(status_code=503, detail="No se pudo consultar el saldo")
    return data


@router.get("/sms/history")
async def sms_history(
    limit: int = 20,
    offset: int = 0,
    current_user: User = Depends(_require_admin),
):
    """Historial de SMS enviados."""
    data = await sms.get_history(limit=limit, offset=offset)
    if data is None:
        raise HTTPException(status_code=503, detail="No se pudo consultar el historial")
    return data


@router.get("/volunteers")
async def list_volunteers(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(_require_admin),
):
    """Lista todos los usuarios excepto víctimas."""
    result = await db.execute(
        select(User).where(
            User.role.in_([UserRole.volunteer, UserRole.coordinator, UserRole.admin]),
        ).order_by(User.created_at.desc())
    )
    users = result.scalars().all()
    return [
        {
            "id": u.id,
            "full_name": u.full_name,
            "last_name": u.last_name or "",
            "role": u.role,
            "phone": u.phone,
            "neighborhood": u.neighborhood or "",
            "skills": u.skills or "",
            "is_active": u.is_active,
            "created_at": u.created_at.isoformat(),
        }
        for u in users
    ]


class UserPatchRequest(BaseModel):
    role: Optional[str] = None
    is_active: Optional[bool] = None


@router.patch("/users/{user_id}")
async def patch_user(
    user_id: int,
    data: UserPatchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(_require_admin),
):
    """Cambia el rol o activa/desactiva un usuario."""
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="No puedes modificarte a ti mismo")

    if data.role is not None:
        try:
            user.role = UserRole(data.role)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Rol inválido: {data.role}")

    if data.is_active is not None:
        user.is_active = data.is_active

    await db.commit()
    await db.refresh(user)
    return {"id": user.id, "role": user.role, "is_active": user.is_active}


@router.post("/run-delivery-migration")
async def run_delivery_migration(current_user: User = Depends(_require_admin)):
    """Ejecuta la migración de medication_deliveries sin reiniciar el servidor."""
    import sqlite3, os, logging
    from pathlib import Path
    log = logging.getLogger(__name__)
    db_path = str(Path(__file__).resolve().parent.parent.parent / "pereira_alerta.db")
    if not os.path.exists(db_path):
        raise HTTPException(404, f"DB no encontrada en {db_path}")
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(medication_deliveries)")
        rows = cur.fetchall()
        if not rows:
            conn.close()
            return {"ok": True, "message": "Tabla no existe aún, se creará al arrancar"}
        cols = {r[1]: r for r in rows}
        added = []
        for col, defn in [
            ("delivery_type",     "TEXT NOT NULL DEFAULT 'solicitude'"),
            ("recipient_id",      "TEXT"),
            ("recipient_phone",   "TEXT"),
            ("recipient_address", "TEXT"),
            ("delivery_group_id", "TEXT"),
        ]:
            if col not in cols:
                cur.execute(f"ALTER TABLE medication_deliveries ADD COLUMN {col} {defn}")
                added.append(col)
                cols[col] = True
        # Recrear si report_id sigue NOT NULL
        cur.execute("PRAGMA table_info(medication_deliveries)")
        info = {r[1]: r for r in cur.fetchall()}
        recreated = False
        if info.get("report_id", (None,)*4)[3]:
            cur.execute("PRAGMA foreign_keys=OFF")
            cur.executescript("""
                CREATE TABLE IF NOT EXISTS med_del_new (
                    id INTEGER PRIMARY KEY,
                    report_id INTEGER REFERENCES reports(id) ON DELETE SET NULL,
                    delivery_type TEXT NOT NULL DEFAULT 'solicitude',
                    delivery_group_id TEXT,
                    stock_id INTEGER REFERENCES medication_stock(id) ON DELETE SET NULL,
                    medication_name VARCHAR(200) NOT NULL,
                    quantity_delivered INTEGER NOT NULL DEFAULT 1,
                    unit VARCHAR(40) NOT NULL DEFAULT 'unidades',
                    delivered_to VARCHAR(150) NOT NULL,
                    recipient_id TEXT,
                    recipient_phone TEXT,
                    recipient_address TEXT,
                    delivered_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    delivery_notes TEXT,
                    delivered_at DATETIME NOT NULL,
                    created_at DATETIME NOT NULL
                );
                INSERT OR IGNORE INTO med_del_new
                    (id,report_id,delivery_type,stock_id,medication_name,quantity_delivered,
                     unit,delivered_to,recipient_id,recipient_phone,recipient_address,
                     delivered_by,delivery_notes,delivered_at,created_at)
                SELECT id,report_id,COALESCE(delivery_type,'solicitude'),stock_id,
                       medication_name,quantity_delivered,unit,delivered_to,
                       recipient_id,recipient_phone,recipient_address,
                       delivered_by,delivery_notes,delivered_at,created_at
                FROM medication_deliveries;
                DROP TABLE medication_deliveries;
                ALTER TABLE med_del_new RENAME TO medication_deliveries;
            """)
            cur.execute("PRAGMA foreign_keys=ON")
            recreated = True
        conn.commit()
        conn.close()
        return {"ok": True, "columns_added": added, "table_recreated": recreated,
                "message": "Migración completada. Ya puedes registrar entregas."}
    except Exception as exc:
        log.error("run_delivery_migration error: %s", exc, exc_info=True)
        raise HTTPException(500, f"Error en migración: {exc}")


@router.get("/stats")
async def system_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(_require_admin),
):
    """Estadísticas globales del sistema."""
    from app.models.report import Report, ReportStatus
    from app.models.danger_zone import DangerZone
    from sqlalchemy import func

    total_users = (await db.execute(select(func.count(User.id)))).scalar()
    total_volunteers = (await db.execute(
        select(func.count(User.id)).where(User.role == UserRole.volunteer)
    )).scalar()
    total_coordinators = (await db.execute(
        select(func.count(User.id)).where(User.role == UserRole.coordinator)
    )).scalar()
    total_reports = (await db.execute(select(func.count(Report.id)))).scalar()
    pending_reports = (await db.execute(
        select(func.count(Report.id)).where(Report.status == ReportStatus.pending)
    )).scalar()
    resolved_reports = (await db.execute(
        select(func.count(Report.id)).where(Report.status == ReportStatus.resolved)
    )).scalar()
    danger_zones = (await db.execute(
        select(func.count(DangerZone.id)).where(DangerZone.is_active == True)  # noqa: E712
    )).scalar()

    return {
        "users": {"total": total_users, "volunteers": total_volunteers, "coordinators": total_coordinators},
        "reports": {"total": total_reports, "pending": pending_reports, "resolved": resolved_reports},
        "danger_zones": danger_zones,
    }
