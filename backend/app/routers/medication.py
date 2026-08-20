"""
Módulo de gestión de medicamentos.

Flujo:
  Ciudadano → Pedir Ayuda (need_type=medicine) → Report creado
  Admin → ve solicitud en /medication/requests
       → agrega detalle del medicamento (nombre exacto, cantidad)
       → registra entrega → Report pasa a resolved
  Admin → gestiona stock (donaciones recibidas)
"""
from __future__ import annotations

import logging
from datetime import datetime
from typing import Optional

import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.medication import MedicationStock, MedicationDelivery
from app.models.report import Report, NeedType, ReportStatus
from app.models.user import User, UserRole
from app.services.auth import get_current_user, get_current_user_from_query

log = logging.getLogger(__name__)
router = APIRouter(prefix="/medication", tags=["Medication"])

# ── helpers ──────────────────────────────────────────────────────────────────

def _stock_payload(s: MedicationStock) -> dict:
    return {
        "id": s.id,
        "name": s.name,
        "quantity": s.quantity,
        "unit": s.unit,
        "expiry_date": s.expiry_date.isoformat() if s.expiry_date else None,
        "storage_location": s.storage_location,
        "aid_point_id": s.aid_point_id,
        "donated_by": s.donated_by,
        "notes": s.notes,
        "is_active": s.is_active,
        "created_at": s.created_at.isoformat() + "Z",
        "updated_at": s.updated_at.isoformat() + "Z",
    }

def _delivery_payload(d: MedicationDelivery) -> dict:
    return {
        "id": d.id,
        "report_id": d.report_id,
        "delivery_type": getattr(d, "delivery_type", "solicitude"),
        "delivery_group_id": getattr(d, "delivery_group_id", None),
        "stock_id": d.stock_id,
        "medication_name": d.medication_name,
        "quantity_delivered": d.quantity_delivered,
        "unit": d.unit,
        "delivered_to": d.delivered_to,
        "recipient_id": getattr(d, "recipient_id", None),
        "recipient_phone": getattr(d, "recipient_phone", None),
        "recipient_address": getattr(d, "recipient_address", None),
        "delivered_by": d.delivered_by,
        "delivery_notes": d.delivery_notes,
        "delivered_at": d.delivered_at.isoformat() + "Z",
        "created_at": d.created_at.isoformat() + "Z",
    }

def _require_admin(user: User):
    if user.role not in (UserRole.admin, UserRole.coordinator):
        raise HTTPException(403, "Sin permisos")

# ── Solicitudes (reports con need_type=medicine) ──────────────────────────────

@router.get("/requests")
async def list_medication_requests(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista de reportes con need_type=medicine, con entregas asociadas."""
    _require_admin(current_user)

    result = await db.execute(
        select(Report)
        .where(Report.need_type == NeedType.medicine)
        .order_by(Report.created_at.desc())
    )
    reports = result.scalars().all()

    out = []
    for r in reports:
        # Entregas de este reporte
        deliveries_res = await db.execute(
            select(MedicationDelivery).where(MedicationDelivery.report_id == r.id)
        )
        deliveries = [_delivery_payload(d) for d in deliveries_res.scalars().all()]

        # Voluntario asignado
        assigned_name = None
        if r.assigned_to:
            u = (await db.execute(select(User).where(User.id == r.assigned_to))).scalar_one_or_none()
            if u:
                assigned_name = f"{u.full_name}{' ' + u.last_name if u.last_name else ''}"

        out.append({
            "id": r.id,
            "reporter_name": r.reporter_name,
            "reporter_phone": r.reporter_phone,
            "description": r.description,
            "address": r.address,
            "lat": r.lat,
            "lng": r.lng,
            "people_count": r.people_count,
            "status": r.status.value,
            "medication_detail": r.medication_detail if hasattr(r, "medication_detail") else None,
            "assigned_to": r.assigned_to,
            "assigned_name": assigned_name,
            "photo_url": r.photo_url,
            "created_at": r.created_at.isoformat() + "Z",
            "updated_at": r.updated_at.isoformat() + "Z",
            "deliveries": deliveries,
        })

    return out


@router.patch("/requests/{report_id}/detail")
async def update_medication_detail(
    report_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Actualiza el detalle de medicamento de una solicitud (nombre exacto, dosis, etc.)."""
    _require_admin(current_user)

    r = (await db.execute(select(Report).where(Report.id == report_id))).scalar_one_or_none()
    if not r or r.need_type != NeedType.medicine:
        raise HTTPException(404, "Solicitud no encontrada")

    detail = body.get("medication_detail", "")
    status = body.get("status")

    update_vals: dict = {"medication_detail": detail, "updated_at": datetime.utcnow()}
    if status and status in ("pending", "in_progress", "resolved"):
        update_vals["status"] = status

    await db.execute(update(Report).where(Report.id == report_id).values(**update_vals))
    await db.commit()
    return {"ok": True}


# ── Stock ─────────────────────────────────────────────────────────────────────

@router.get("/stock")
async def list_stock(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista el inventario de medicamentos disponibles."""
    _require_admin(current_user)
    result = await db.execute(
        select(MedicationStock)
        .where(MedicationStock.is_active == True)  # noqa
        .order_by(MedicationStock.name)
    )
    return [_stock_payload(s) for s in result.scalars().all()]


@router.post("/stock")
async def create_stock(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Registra un nuevo medicamento en el inventario."""
    _require_admin(current_user)

    name = str(body.get("name", "")).strip()
    if not name:
        raise HTTPException(400, "El nombre es requerido")

    expiry = None
    if body.get("expiry_date"):
        try:
            from datetime import date
            expiry = date.fromisoformat(body["expiry_date"])
        except Exception:
            pass

    s = MedicationStock(
        name=name,
        quantity=int(body.get("quantity", 0)),
        unit=str(body.get("unit", "unidades")).strip(),
        expiry_date=expiry,
        storage_location=body.get("storage_location"),
        aid_point_id=body.get("aid_point_id"),
        donated_by=body.get("donated_by"),
        notes=body.get("notes"),
        created_by=current_user.id,
    )
    db.add(s)
    await db.commit()
    await db.refresh(s)
    return _stock_payload(s)


@router.patch("/stock/{stock_id}")
async def update_stock(
    stock_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Actualiza cantidad, ubicación u otros datos del stock."""
    _require_admin(current_user)

    s = (await db.execute(select(MedicationStock).where(MedicationStock.id == stock_id))).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Stock no encontrado")

    allowed = {"name", "quantity", "unit", "storage_location", "donated_by", "notes", "is_active"}
    vals: dict = {"updated_at": datetime.utcnow()}
    for k, v in body.items():
        if k in allowed:
            vals[k] = v

    if "expiry_date" in body:
        try:
            from datetime import date
            vals["expiry_date"] = date.fromisoformat(body["expiry_date"]) if body["expiry_date"] else None
        except Exception:
            pass

    await db.execute(update(MedicationStock).where(MedicationStock.id == stock_id).values(**vals))
    await db.commit()
    return {"ok": True}


@router.delete("/stock/{stock_id}")
async def deactivate_stock(
    stock_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Desactiva un ítem del inventario (soft delete)."""
    _require_admin(current_user)
    await db.execute(
        update(MedicationStock)
        .where(MedicationStock.id == stock_id)
        .values(is_active=False, updated_at=datetime.utcnow())
    )
    await db.commit()
    return {"ok": True}


# ── Entregas ──────────────────────────────────────────────────────────────────

@router.get("/deliveries")
async def list_deliveries(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista todas las entregas registradas."""
    _require_admin(current_user)
    result = await db.execute(
        select(MedicationDelivery).order_by(MedicationDelivery.delivered_at.desc())
    )
    return [_delivery_payload(d) for d in result.scalars().all()]


@router.post("/deliveries")
async def create_delivery(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Registra la entrega de un medicamento a un solicitante.
    Si viene stock_id, descuenta la cantidad del inventario.
    Opcionalmente marca el reporte como resuelto.
    """
    _require_admin(current_user)

    report_id = body.get("report_id") or None
    delivery_type = "direct" if not report_id else "solicitude"
    qty = int(body.get("quantity_delivered", 1))

    # Validar reporte si viene con report_id
    r = None
    if report_id:
        r = (await db.execute(select(Report).where(Report.id == report_id))).scalar_one_or_none()
        if not r:
            raise HTTPException(404, "Reporte no encontrado")

    # Descontar del stock si se indicó
    stock_id = body.get("stock_id") or None
    if stock_id:
        s = (await db.execute(select(MedicationStock).where(MedicationStock.id == stock_id))).scalar_one_or_none()
        if s:
            new_qty = max(0, s.quantity - qty)
            await db.execute(
                update(MedicationStock)
                .where(MedicationStock.id == stock_id)
                .values(quantity=new_qty, updated_at=datetime.utcnow())
            )

    delivered_to = str(body.get("delivered_to", "") or (r.reporter_name if r else "")).strip() or "Sin nombre"

    d = MedicationDelivery(
        report_id=report_id,
        delivery_type=delivery_type,
        delivery_group_id=body.get("delivery_group_id") or None,
        stock_id=stock_id,
        medication_name=str(body.get("medication_name", "")).strip() or "Sin especificar",
        quantity_delivered=qty,
        unit=str(body.get("unit", "unidades")).strip(),
        delivered_to=delivered_to,
        recipient_id=body.get("recipient_id") or None,
        recipient_phone=body.get("recipient_phone") or None,
        recipient_address=body.get("recipient_address") or None,
        delivered_by=current_user.id,
        delivery_notes=body.get("delivery_notes"),
        delivered_at=datetime.utcnow(),
    )
    db.add(d)

    # Marcar reporte como resuelto si se solicita
    if body.get("resolve_report") and report_id:
        await db.execute(
            update(Report)
            .where(Report.id == report_id)
            .values(status=ReportStatus.resolved, updated_at=datetime.utcnow())
        )

    try:
        await db.commit()
        await db.refresh(d)
    except Exception as exc:
        await db.rollback()
        log.error("create_delivery DB error: %s", exc, exc_info=True)
        raise HTTPException(500, f"Error al guardar entrega: {exc}")

    return _delivery_payload(d)


@router.get("/db-schema-check")
async def db_schema_check(current_user: User = Depends(get_current_user)):
    """Diagnóstico temporal: muestra columnas actuales de medication_deliveries."""
    _require_admin(current_user)
    import sqlite3
    from pathlib import Path as P
    db_path = str(P(__file__).resolve().parent.parent.parent / "pereira_alerta.db")
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(medication_deliveries)")
        rows = cur.fetchall()
        conn.close()
        return {
            "db_path": db_path,
            "columns": [{"name": r[1], "type": r[2], "notnull": bool(r[3]), "default": r[4]} for r in rows],
        }
    except Exception as exc:
        return {"error": str(exc), "db_path": db_path}


# ── Exportación Excel ─────────────────────────────────────────────────────────

@router.get("/stock/export")
async def export_stock_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_query),
):
    """Exporta el inventario de medicamentos activo como archivo Excel."""
    _require_admin(current_user)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    result = await db.execute(
        select(MedicationStock)
        .where(MedicationStock.is_active == True)  # noqa
        .order_by(MedicationStock.name)
    )
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Medicamentos"

    # ── Encabezado principal ──────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "Te Ayudo Pereira — Inventario de Medicamentos"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="DC2626")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    date_cell = ws["A2"]
    date_cell.value = f"Generado el {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC"
    date_cell.font = Font(italic=True, size=9, color="6B7280")
    date_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # ── Cabeceras de columnas ─────────────────────────────────────────────
    headers = ["Medicamento", "Cantidad", "Unidad", "Vencimiento", "Almacenamiento", "Donado por", "Fecha registro"]
    header_fill = PatternFill("solid", fgColor="111827")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 20

    # ── Filas de datos ────────────────────────────────────────────────────
    red_fill   = PatternFill("solid", fgColor="FEF2F2")  # qty 0
    yellow_fill = PatternFill("solid", fgColor="FFFBEB")  # qty < 5
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, s in enumerate(items, start=4):
        qty = s.quantity
        row_fill = red_fill if qty == 0 else (yellow_fill if qty < 5 else white_fill)

        values = [
            s.name,
            qty,
            s.unit,
            s.expiry_date.strftime("%d/%m/%Y") if s.expiry_date else "—",
            s.storage_location or "—",
            s.donated_by or "—",
            s.created_at.strftime("%d/%m/%Y") if s.created_at else "—",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx == 2:  # Cantidad — centrar y negrita
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True,
                                 color="DC2626" if qty == 0 else ("D97706" if qty < 5 else "16A34A"))

    # ── Leyenda ───────────────────────────────────────────────────────────
    legend_row = len(items) + 5
    ws.merge_cells(f"A{legend_row}:G{legend_row}")
    legend = ws[f"A{legend_row}"]
    legend.value = "🔴 Sin stock   🟡 Stock bajo (< 5)   🟢 Disponible"
    legend.font = Font(size=9, color="6B7280", italic=True)
    legend.alignment = Alignment(horizontal="left")

    # ── Anchos de columna ─────────────────────────────────────────────────
    col_widths = [36, 12, 14, 16, 28, 24, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Congelar encabezados ──────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Serializar y devolver ─────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"inventario_medicamentos_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
